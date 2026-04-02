from typing import Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def export_excel(registros: List[Dict], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    if not registros:
        wb.save(filepath)
        return

    headers = list(registros[0].keys())
    ws.append(headers)
    for row in registros:
        ws.append([row.get(h) for h in headers])

    # largura básica
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(str(header)) + 2, 12), 32)

    wb.save(filepath)
